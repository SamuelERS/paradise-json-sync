"""
Tests for PurchaseExporter / Tests del Exportador de Compras
=============================================================

Unit tests for the configurable purchase invoice exporter.
Tests unitarios para el exportador configurable de facturas de compra.
"""

import csv
import json
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.purchases.purchase_exporter import (
    ALL_COLUMNS,
    CURRENCY_COLUMNS,
    PROFILE_BASICO,
    PROFILE_COMPLETO,
    PROFILE_CONTADOR,
    PurchaseExporter,
)
from src.models.purchase_invoice import (
    PurchaseDocumentType,
    PurchaseInvoice,
    PurchaseInvoiceItem,
    SupplierInfo,
)


@pytest.fixture
def exporter() -> PurchaseExporter:
    """Fixture for PurchaseExporter instance."""
    return PurchaseExporter()


@pytest.fixture
def second_invoice() -> PurchaseInvoice:
    """Second invoice for multi-invoice tests."""
    return PurchaseInvoice(
        document_number="X9Y8Z7W6-V5U4-3210-FEDC-BA9876543210",
        control_number="DTE-01-00000002-000000000000002",
        document_type=PurchaseDocumentType.FACTURA,
        issue_date=date(2026, 2, 7),
        emission_time="09:00:00",
        currency="USD",
        supplier=SupplierInfo(
            name="PROVEEDOR XYZ S.A.",
            nit="0614-654321-987-0",
        ),
        receiver_name="MI EMPRESA S.A. DE C.V.",
        receiver_nit="0614-999999-999-9",
        items=[
            PurchaseInvoiceItem(
                item_number=1,
                description="Toner HP LaserJet",
                quantity=Decimal("2"),
                unit_price=Decimal("45.00"),
                taxable_sale=Decimal("90.00"),
                total=Decimal("90.00"),
            ),
        ],
        subtotal=Decimal("90.00"),
        total_taxable=Decimal("90.00"),
        tax=Decimal("11.70"),
        total=Decimal("101.70"),
        payment_condition=2,
        source_file="factura_xyz_002.json",
        detected_format="DTE_STANDARD",
        detection_confidence=0.88,
    )


# === TestColumnResolution ===


class TestColumnResolution:
    """Tests for column profile resolution. / Tests de resolucion de perfiles."""

    def test_resolve_basico_profile(self, exporter: PurchaseExporter) -> None:
        """Basic profile resolves to 10 columns."""
        cols = exporter._resolve_columns("basico", None)
        assert len(cols) == 10
        assert [c.id for c in cols] == PROFILE_BASICO

    def test_resolve_completo_profile(self, exporter: PurchaseExporter) -> None:
        """Complete profile resolves to all 32 columns."""
        cols = exporter._resolve_columns("completo", None)
        assert len(cols) == 32
        assert len(cols) == len(ALL_COLUMNS)

    def test_resolve_contador_profile(self, exporter: PurchaseExporter) -> None:
        """Accountant profile resolves to 15 columns."""
        cols = exporter._resolve_columns("contador", None)
        assert len(cols) == 15
        assert [c.id for c in cols] == PROFILE_CONTADOR

    def test_resolve_custom_columns(self, exporter: PurchaseExporter) -> None:
        """Custom profile uses only selected columns."""
        custom = ["supplier_name", "total", "issue_date"]
        cols = exporter._resolve_columns("custom", custom)
        assert len(cols) == 3
        assert [c.id for c in cols] == custom

    def test_resolve_invalid_column(self, exporter: PurchaseExporter) -> None:
        """Unknown columns are silently ignored."""
        custom = ["supplier_name", "NONEXISTENT", "total"]
        cols = exporter._resolve_columns("custom", custom)
        assert len(cols) == 2
        assert [c.id for c in cols] == ["supplier_name", "total"]


# === TestColumnValueMapping ===


class TestColumnValueMapping:
    """Tests for column value extraction. / Tests de extraccion de valores."""

    def test_get_all_column_values(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice
    ) -> None:
        """Every column returns a value for a complete invoice."""
        inv = sample_purchase_invoice
        for col in ALL_COLUMNS:
            value = exporter._get_column_value(inv, col.id)
            assert value is not None, f"Column {col.id} returned None"

    def test_currency_columns_return_float(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice
    ) -> None:
        """Currency columns return float values, not Decimal."""
        inv = sample_purchase_invoice
        for col_id in CURRENCY_COLUMNS:
            value = exporter._get_column_value(inv, col_id)
            assert isinstance(value, float), f"{col_id} returned {type(value)}, expected float"

    def test_payment_condition_text(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice
    ) -> None:
        """Payment condition: 1=CONTADO, 2=CRÉDITO, None=empty."""
        inv = sample_purchase_invoice
        assert exporter._get_column_value(inv, "payment_condition") == "CONTADO"

        inv_credit = inv.model_copy(update={"payment_condition": 2})
        assert exporter._get_column_value(inv_credit, "payment_condition") == "CRÉDITO"

        inv_none = inv.model_copy(update={"payment_condition": None})
        assert exporter._get_column_value(inv_none, "payment_condition") == ""


# === TestExcelExport ===


class TestExcelExport:
    """Tests for Excel export. / Tests de exportacion Excel."""

    def test_export_excel_basico(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """Excel with basic profile generates 10 columns."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path), format="xlsx", column_profile="basico",
        )
        assert Path(path).exists()
        wb = load_workbook(path)
        sheet = wb.active
        assert sheet.max_column == 10
        assert sheet.max_row == 2  # header + 1 invoice

    def test_export_excel_with_summary(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """Excel with include_summary creates Resumen sheet."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path), format="xlsx",
            options={"include_summary": True},
        )
        wb = load_workbook(path)
        assert "Resumen" in wb.sheetnames
        summary = wb["Resumen"]
        assert summary.cell(row=1, column=1).value == "Proveedor"
        # Check TOTAL row exists
        last_row = summary.max_row
        assert summary.cell(row=last_row, column=1).value == "TOTAL"

    def test_export_excel_with_items(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """Excel with include_items_sheet creates Items sheet."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path), format="xlsx",
            options={"include_items_sheet": True},
        )
        wb = load_workbook(path)
        assert "Items" in wb.sheetnames
        items_sheet = wb["Items"]
        assert items_sheet.cell(row=1, column=4).value == "Descripción"
        assert items_sheet.max_row == 2  # header + 1 item

    def test_export_excel_currency_format(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """Currency cells use $#,##0.00 format."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path), format="xlsx", column_profile="basico",
        )
        wb = load_workbook(path)
        sheet = wb.active
        # In basico profile: subtotal=col6, tax=col7, total=col8
        for col_idx in [6, 7, 8]:
            cell = sheet.cell(row=2, column=col_idx)
            assert "$#,##0.00" in cell.number_format


# === TestCSVExport ===


class TestCSVExport:
    """Tests for CSV export. / Tests de exportacion CSV."""

    def test_export_csv_basic(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """CSV with basic profile generates correct headers and data."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path), format="csv", column_profile="basico",
        )
        assert Path(path).exists()
        with open(path, encoding="utf-8") as f:
            reader = list(csv.reader(f))
        assert len(reader) == 2  # header + 1 row
        assert len(reader[0]) == 10
        assert reader[0][0] == "N° Control"


# === TestJSONExport ===


class TestJSONExport:
    """Tests for JSON export. / Tests de exportacion JSON."""

    def test_export_json_always_complete(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """JSON always includes ALL 32 column fields regardless of profile."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path),
            format="json", column_profile="basico",
        )
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        inv_data = data["invoices"][0]
        for col in ALL_COLUMNS:
            assert col.id in inv_data, f"Missing field: {col.id}"
        assert "items" in inv_data

    def test_export_json_with_raw_data(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """JSON with include_raw_data=True includes raw_data."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path),
            format="json", options={"include_raw_data": True},
        )
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        assert "raw_data" in data["invoices"][0]

    def test_export_json_metadata(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """JSON includes metadata with exported_at, total_invoices, etc."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path), format="json",
        )
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        meta = data["metadata"]
        assert "exported_at" in meta
        assert meta["total_invoices"] == 1
        assert meta["total_amount"] == pytest.approx(39.55)
        assert meta["format_version"] == "1.0"


# === TestPDFExport ===


class TestPDFExport:
    """Tests for PDF export. / Tests de exportacion PDF."""

    def test_export_pdf_basic(
        self, exporter: PurchaseExporter, sample_purchase_invoice: PurchaseInvoice, tmp_path: Path,
    ) -> None:
        """PDF export generates a valid file with size > 0."""
        path = exporter.export(
            [sample_purchase_invoice], str(tmp_path),
            format="pdf", column_profile="basico",
        )
        p = Path(path)
        assert p.exists()
        assert p.stat().st_size > 0
        assert p.suffix == ".pdf"


# === TestEdgeCases ===


class TestEdgeCases:
    """Tests for edge cases. / Tests de casos borde."""

    def test_empty_invoices_list(self, exporter: PurchaseExporter, tmp_path: Path) -> None:
        """Empty invoice list generates file with only headers (no error)."""
        path = exporter.export([], str(tmp_path), format="xlsx", column_profile="basico")
        wb = load_workbook(path)
        sheet = wb.active
        assert sheet.max_row == 1  # only headers
        assert sheet.max_column == 10

    def test_group_by_supplier(
        self,
        exporter: PurchaseExporter,
        sample_purchase_invoice: PurchaseInvoice,
        second_invoice: PurchaseInvoice,
        tmp_path: Path,
    ) -> None:
        """Rows are grouped by supplier name in CSV."""
        path = exporter.export(
            [second_invoice, sample_purchase_invoice],
            str(tmp_path), format="csv", column_profile="basico",
            options={"group_by": "supplier"},
        )
        with open(path, encoding="utf-8") as f:
            rows = list(csv.reader(f))
        # Headers + 2 data rows; DISTRIBUIDORA < PROVEEDOR alphabetically
        suppliers = [rows[1][3], rows[2][3]]  # supplier_name is col index 3 in basico
        assert suppliers == sorted(suppliers)
