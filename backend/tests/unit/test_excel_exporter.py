"""
ExcelExporter Tests / Pruebas del Exportador de Excel
=====================================================

Unit tests for the ExcelExporter class.
Pruebas unitarias para la clase ExcelExporter.
"""

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.core.excel_exporter import ExcelExporter, ExcelExporterError


class TestExcelExporterInit:
    """Tests for ExcelExporter initialization."""

    def test_default_initialization(self):
        """Test default initialization values."""
        exporter = ExcelExporter()
        assert exporter.currency_symbol == "$"
        assert exporter.decimal_places == 2

    def test_custom_initialization(self):
        """Test custom initialization values."""
        exporter = ExcelExporter(currency_symbol="€", decimal_places=4)
        assert exporter.currency_symbol == "€"
        assert exporter.decimal_places == 4


class TestExportToExcel:
    """Tests for export_to_excel method."""

    def test_export_single_invoice(self, sample_invoice, temp_output_dir):
        """Test exporting a single invoice."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "test_export.xlsx")

        result = exporter.export_to_excel([sample_invoice], output_path)

        assert Path(result).exists()
        assert result.endswith(".xlsx")

    def test_export_multiple_invoices(self, multiple_invoices, temp_output_dir):
        """Test exporting multiple invoices."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "test_multiple.xlsx")

        result = exporter.export_to_excel(multiple_invoices, output_path)

        assert Path(result).exists()

        # Verify content
        wb = load_workbook(result)
        ws = wb.active
        # Should have header row + 3 data rows
        assert ws.max_row == 4

    def test_export_creates_output_directory(self, sample_invoice, tmp_path):
        """Test that export creates output directory if needed."""
        exporter = ExcelExporter()
        output_path = str(tmp_path / "new_dir" / "subdir" / "export.xlsx")

        result = exporter.export_to_excel([sample_invoice], output_path)

        assert Path(result).exists()

    def test_export_empty_list_raises_error(self, temp_output_dir):
        """Test that exporting empty list raises error."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "empty.xlsx")

        with pytest.raises(ExcelExporterError) as exc_info:
            exporter.export_to_excel([], output_path)

        assert "no invoice" in exc_info.value.message.lower()

    def test_export_with_summary_sheet(self, multiple_invoices, temp_output_dir):
        """Test that summary sheet is created when requested."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "with_summary.xlsx")

        exporter.export_to_excel(multiple_invoices, output_path, include_summary=True)

        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Summary - Resumen" in sheet_names

    def test_export_without_summary_sheet(self, multiple_invoices, temp_output_dir):
        """Test that summary sheet is not created when not requested."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "no_summary.xlsx")

        exporter.export_to_excel(multiple_invoices, output_path, include_summary=False)

        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Summary - Resumen" not in sheet_names

    def test_export_with_items_sheet(self, multiple_invoices, temp_output_dir):
        """Test that items sheet is created when requested."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "with_items.xlsx")

        exporter.export_to_excel(multiple_invoices, output_path, include_items=True)

        wb = load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert "Items - Items" in sheet_names


class TestExportHeaders:
    """Tests for Excel headers."""

    def test_correct_headers(self, sample_invoice, temp_output_dir):
        """Test that correct headers are created."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "headers_test.xlsx")

        exporter.export_to_excel([sample_invoice], output_path, include_summary=False)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check headers
        headers = [cell.value for cell in ws[1]]
        expected_headers = [
            "Document Number / Número",
            "Type / Tipo",
            "Issue Date / Fecha",
            "Customer Name / Cliente",
            "Customer ID / NIT",
            "Subtotal",
            "Tax / Impuesto",
            "Total",
            "Source File / Archivo",
        ]

        assert headers == expected_headers

    def test_headers_are_styled(self, sample_invoice, temp_output_dir):
        """Test that headers have styling applied."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "styled_headers.xlsx")

        exporter.export_to_excel([sample_invoice], output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Check first header cell styling
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True


class TestFormatCurrency:
    """Tests for format_currency method."""

    def test_format_with_default_symbol(self):
        """Test currency formatting with default symbol."""
        exporter = ExcelExporter()
        result = exporter.format_currency(Decimal("1234.56"))

        assert result == "$1,234.56"

    def test_format_with_custom_symbol(self):
        """Test currency formatting with custom symbol."""
        exporter = ExcelExporter(currency_symbol="€")
        result = exporter.format_currency(Decimal("1234.56"))

        assert result == "€1,234.56"

    def test_format_with_more_decimals(self):
        """Test currency formatting with more decimal places."""
        exporter = ExcelExporter(decimal_places=4)
        result = exporter.format_currency(Decimal("1234.5678"))

        assert result == "$1,234.5678"

    def test_format_zero(self):
        """Test formatting zero value."""
        exporter = ExcelExporter()
        result = exporter.format_currency(Decimal("0"))

        assert result == "$0.00"

    def test_format_large_number(self):
        """Test formatting large number with thousands separator."""
        exporter = ExcelExporter()
        result = exporter.format_currency(Decimal("1234567.89"))

        assert result == "$1,234,567.89"


class TestSummarySheet:
    """Tests for summary sheet creation."""

    def test_summary_contains_totals(self, multiple_invoices, temp_output_dir):
        """Test that summary sheet contains correct totals."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "summary_totals.xlsx")

        exporter.export_to_excel(multiple_invoices, output_path, include_summary=True)

        wb = load_workbook(output_path)
        summary_sheet = wb["Summary - Resumen"]

        # Find totals in the sheet
        sheet_values = [
            str(cell.value) for row in summary_sheet.iter_rows() for cell in row if cell.value
        ]

        # Should contain total count
        assert "3" in sheet_values  # 3 invoices

    def test_summary_contains_type_breakdown(self, multiple_invoices, temp_output_dir):
        """Test that summary contains invoice type breakdown."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "summary_types.xlsx")

        exporter.export_to_excel(multiple_invoices, output_path, include_summary=True)

        wb = load_workbook(output_path)
        summary_sheet = wb["Summary - Resumen"]

        sheet_text = " ".join(
            str(cell.value) for row in summary_sheet.iter_rows() for cell in row if cell.value
        )

        # Should mention invoice types
        assert "Factura" in sheet_text
        assert "CCF" in sheet_text


class TestItemsSheet:
    """Tests for items detail sheet."""

    def test_items_sheet_created(self, multiple_invoices, temp_output_dir):
        """Test that items sheet is created with correct data."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "items_sheet.xlsx")

        exporter.export_to_excel(multiple_invoices, output_path, include_items=True)

        wb = load_workbook(output_path)
        items_sheet = wb["Items - Items"]

        # Should have header + items (one item per invoice)
        assert items_sheet.max_row >= 4  # Header + 3 items (one per invoice)

    def test_items_sheet_headers(self, sample_invoice, temp_output_dir):
        """Test that items sheet has correct headers."""
        exporter = ExcelExporter()
        output_path = str(temp_output_dir / "items_headers.xlsx")

        exporter.export_to_excel([sample_invoice], output_path, include_items=True)

        wb = load_workbook(output_path)
        items_sheet = wb["Items - Items"]

        headers = [cell.value for cell in items_sheet[1]]

        assert "Invoice / Factura" in headers
        assert "Description / Descripción" in headers
        assert "Total" in headers


class TestExcelExporterError:
    """Tests for ExcelExporterError exception."""

    def test_error_with_message_only(self):
        """Test creating error with message only."""
        error = ExcelExporterError("Test error")
        assert str(error) == "Test error"
        assert error.output_path is None

    def test_error_with_output_path(self):
        """Test creating error with output path."""
        error = ExcelExporterError(
            message="Export failed",
            output_path="/path/to/output.xlsx",
        )
        assert error.output_path == "/path/to/output.xlsx"

    def test_error_with_original_error(self):
        """Test creating error with original exception."""
        original = OSError("Disk full")
        error = ExcelExporterError(
            message="Export failed",
            original_error=original,
        )
        assert error.original_error is original
