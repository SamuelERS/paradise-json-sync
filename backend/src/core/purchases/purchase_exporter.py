"""
Purchase Exporter / Exportador de Facturas de Compra
=====================================================

Configurable exporter for purchase invoices (Excel, CSV, PDF, JSON).
Exportador configurable de facturas de compra (Excel, CSV, PDF, JSON).
"""

import csv
import json
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.purchase_invoice import PurchaseInvoice

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class ColumnDef:
    """Column definition. / Definicion de columna."""

    id: str
    label: str
    category: str
    is_currency: bool = False


# === ALL COLUMNS / TODAS LAS COLUMNAS (32 total, 5 categorias) ===

ALL_COLUMNS: list[ColumnDef] = [
    # Identificacion (7)
    ColumnDef("control_number", "N° Control", "Identificación"),
    ColumnDef("document_number", "Código Gen.", "Identificación"),
    ColumnDef("document_type", "Tipo Doc", "Identificación"),
    ColumnDef("issue_date", "Fecha", "Identificación"),
    ColumnDef("emission_time", "Hora", "Identificación"),
    ColumnDef("currency", "Moneda", "Identificación"),
    ColumnDef("dte_version", "Versión DTE", "Identificación"),
    # Proveedor (8)
    ColumnDef("supplier_name", "Proveedor", "Proveedor"),
    ColumnDef("supplier_commercial", "Nombre Comercial", "Proveedor"),
    ColumnDef("supplier_nit", "NIT Proveedor", "Proveedor"),
    ColumnDef("supplier_nrc", "NRC Proveedor", "Proveedor"),
    ColumnDef("supplier_activity", "Actividad", "Proveedor"),
    ColumnDef("supplier_address", "Dir. Proveedor", "Proveedor"),
    ColumnDef("supplier_phone", "Tel. Proveedor", "Proveedor"),
    ColumnDef("supplier_email", "Email Proveedor", "Proveedor"),
    # Receptor (3)
    ColumnDef("receiver_name", "Receptor", "Receptor"),
    ColumnDef("receiver_nit", "NIT Receptor", "Receptor"),
    ColumnDef("receiver_nrc", "NRC Receptor", "Receptor"),
    # Montos (9)
    ColumnDef("total_taxable", "Gravado", "Montos", is_currency=True),
    ColumnDef("total_exempt", "Exento", "Montos", is_currency=True),
    ColumnDef("total_non_subject", "No Sujeto", "Montos", is_currency=True),
    ColumnDef("total_discount", "Descuento", "Montos", is_currency=True),
    ColumnDef("subtotal", "Subtotal", "Montos", is_currency=True),
    ColumnDef("tax", "IVA", "Montos", is_currency=True),
    ColumnDef("iva_retained", "IVA Retenido", "Montos", is_currency=True),
    ColumnDef("total", "Total", "Montos", is_currency=True),
    ColumnDef("total_in_words", "Total Letras", "Montos"),
    # Pago y Metadatos (5)
    ColumnDef("payment_condition", "Condición", "Pago y Metadatos"),
    ColumnDef("tax_seal", "Sello Fiscal", "Pago y Metadatos"),
    ColumnDef("source_file", "Archivo", "Pago y Metadatos"),
    ColumnDef("detected_format", "Formato", "Pago y Metadatos"),
    ColumnDef("detection_confidence", "Confianza", "Pago y Metadatos"),
]

_COLUMN_MAP: dict[str, ColumnDef] = {col.id: col for col in ALL_COLUMNS}
CURRENCY_COLUMNS: set[str] = {col.id for col in ALL_COLUMNS if col.is_currency}

# === PERFILES PREDETERMINADOS ===

PROFILE_BASICO: list[str] = [
    "control_number", "document_type", "issue_date",
    "supplier_name", "supplier_nit",
    "subtotal", "tax", "total",
    "payment_condition", "source_file",
]

PROFILE_COMPLETO: list[str] = [col.id for col in ALL_COLUMNS]

PROFILE_CONTADOR: list[str] = [
    "control_number", "document_type", "issue_date", "emission_time",
    "supplier_name", "supplier_nit", "supplier_nrc", "receiver_nit",
    "total_taxable", "total_exempt", "total_non_subject",
    "total_discount", "subtotal", "tax", "total",
]

_PROFILES: dict[str, list[str]] = {
    "basico": PROFILE_BASICO,
    "completo": PROFILE_COMPLETO,
    "contador": PROFILE_CONTADOR,
}


class PurchaseExporter:
    """
    Configurable purchase invoice exporter (Excel, CSV, PDF, JSON).
    Exportador configurable de facturas de compra.
    """

    def __init__(self, currency_symbol: str = "$", decimal_places: int = 2) -> None:
        """Initialize exporter. / Inicializa el exportador."""
        self.currency_symbol = currency_symbol
        self.decimal_places = decimal_places

    def export(
        self,
        invoices: list[PurchaseInvoice],
        output_dir: str,
        format: str = "xlsx",
        column_profile: str = "completo",
        custom_columns: Optional[list[str]] = None,
        options: Optional[dict[str, Any]] = None,
    ) -> str:
        """Export invoices, return path. / Exporta facturas, retorna ruta."""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        columns = self._resolve_columns(column_profile, custom_columns)
        opts = options or {}
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = str(Path(output_dir) / f"compras_{ts}.{format}")
        if format == "json":
            return self._export_json(invoices, output_path, opts)
        dispatchers = {"xlsx": self._export_excel, "csv": self._export_csv, "pdf": self._export_pdf}
        exporter = dispatchers.get(format, self._export_excel)
        return exporter(invoices, columns, output_path, opts)

    def _resolve_columns(self, profile: str, custom: Optional[list[str]]) -> list[ColumnDef]:
        """Resolve columns from profile or custom list. / Resuelve columnas."""
        if profile == "custom" and custom:
            col_ids = [c for c in custom if c in _COLUMN_MAP]
        else:
            col_ids = _PROFILES.get(profile, PROFILE_COMPLETO)
        return [_COLUMN_MAP[cid] for cid in col_ids]

    def _get_column_value(self, invoice: PurchaseInvoice, column_id: str) -> Any:
        """Extract value from invoice for a column. / Extrae valor de columna."""
        inv = invoice
        if column_id == "payment_condition":
            return "CONTADO" if inv.payment_condition == 1 else "CRÉDITO" if inv.payment_condition == 2 else ""
        if column_id == "detection_confidence":
            return f"{(inv.detection_confidence or 0) * 100:.0f}%"
        mapping: dict[str, Any] = {
            "control_number": inv.control_number or "",
            "document_number": inv.document_number,
            "document_type": inv.document_type.value,
            "issue_date": inv.issue_date.isoformat(),
            "emission_time": inv.emission_time or "",
            "currency": inv.currency,
            "dte_version": str(inv.dte_version) if inv.dte_version else "",
            "supplier_name": inv.supplier.name,
            "supplier_commercial": inv.supplier.commercial_name or "",
            "supplier_nit": inv.supplier.nit or "",
            "supplier_nrc": inv.supplier.nrc or "",
            "supplier_activity": inv.supplier.economic_activity or "",
            "supplier_address": inv.supplier.address or "",
            "supplier_phone": inv.supplier.phone or "",
            "supplier_email": inv.supplier.email or "",
            "receiver_name": inv.receiver_name or "",
            "receiver_nit": inv.receiver_nit or "",
            "receiver_nrc": inv.receiver_nrc or "",
            "total_taxable": float(inv.total_taxable),
            "total_exempt": float(inv.total_exempt),
            "total_non_subject": float(inv.total_non_subject),
            "total_discount": float(inv.total_discount),
            "subtotal": float(inv.subtotal),
            "tax": float(inv.tax),
            "iva_retained": float(inv.iva_retained),
            "total": float(inv.total),
            "total_in_words": inv.total_in_words or "",
            "tax_seal": inv.tax_seal or "",
            "source_file": inv.source_file or "",
            "detected_format": inv.detected_format or "",
        }
        return mapping.get(column_id, "")

    def _export_excel(
        self, invoices: list[PurchaseInvoice], columns: list[ColumnDef],
        output_path: str, options: dict[str, Any],
    ) -> str:
        """Export to Excel with styled headers. / Exporta a Excel."""
        wb = Workbook()
        self._write_main_sheet(wb, invoices, columns)
        if options.get("include_summary"):
            self._create_summary_sheet(wb, invoices)
        if options.get("include_items_sheet"):
            self._create_items_sheet(wb, invoices)
        wb.save(output_path)
        logger.info("Exported %d invoices to Excel: %s", len(invoices), output_path)
        return output_path

    def _write_main_sheet(
        self, wb: Workbook, invoices: list[PurchaseInvoice], columns: list[ColumnDef],
    ) -> None:
        """Write main data sheet. / Escribe hoja principal."""
        sheet = wb.active
        sheet.title = "Compras"
        for ci, col in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=ci, value=col.label)
            self._style_header_cell(cell)
        for ri, inv in enumerate(invoices, 2):
            for ci, col in enumerate(columns, 1):
                val = self._get_column_value(inv, col.id)
                cell = sheet.cell(row=ri, column=ci, value=val)
                if col.is_currency:
                    cell.number_format = f"{self.currency_symbol}#,##0.00"
        self._auto_adjust_widths(sheet)

    def _create_summary_sheet(self, workbook: Workbook, invoices: list[PurchaseInvoice]) -> None:
        """Create summary sheet by supplier. / Crea hoja resumen por proveedor."""
        sheet = workbook.create_sheet(title="Resumen")
        headers = ["Proveedor", "N° Facturas", "Total Gravado", "Total IVA", "Total General"]
        for ci, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=ci, value=h)
            self._style_header_cell(cell)
        supplier_data = self._aggregate_by_supplier(invoices)
        row = 2
        g_count, g_taxable, g_tax, g_total = 0, 0.0, 0.0, 0.0
        for name, d in sorted(supplier_data.items()):
            sheet.cell(row=row, column=1, value=name)
            sheet.cell(row=row, column=2, value=d["count"])
            for c, key in [(3, "taxable"), (4, "tax"), (5, "total")]:
                cell = sheet.cell(row=row, column=c, value=d[key])
                cell.number_format = f"{self.currency_symbol}#,##0.00"
            g_count += d["count"]; g_taxable += d["taxable"]
            g_tax += d["tax"]; g_total += d["total"]
            row += 1
        self._write_summary_totals(sheet, row, g_count, g_taxable, g_tax, g_total)
        self._auto_adjust_widths(sheet)

    def _write_summary_totals(
        self, sheet: Any, row: int, count: int, taxable: float, tax: float, total: float,
    ) -> None:
        """Write TOTAL row in summary. / Escribe fila TOTAL."""
        sheet.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        sheet.cell(row=row, column=2, value=count)
        for c, val in [(3, taxable), (4, tax), (5, total)]:
            cell = sheet.cell(row=row, column=c, value=val)
            cell.number_format = f"{self.currency_symbol}#,##0.00"
            cell.font = Font(bold=True)

    def _aggregate_by_supplier(self, invoices: list[PurchaseInvoice]) -> dict[str, dict[str, Any]]:
        """Aggregate data by supplier. / Agrega datos por proveedor."""
        result: dict[str, dict[str, Any]] = {}
        for inv in invoices:
            name = inv.supplier.name
            if name not in result:
                result[name] = {"count": 0, "taxable": 0.0, "tax": 0.0, "total": 0.0}
            result[name]["count"] += 1
            result[name]["taxable"] += float(inv.total_taxable)
            result[name]["tax"] += float(inv.tax)
            result[name]["total"] += float(inv.total)
        return result

    def _create_items_sheet(self, workbook: Workbook, invoices: list[PurchaseInvoice]) -> None:
        """Create items detail sheet. / Crea hoja de detalle de items."""
        sheet = workbook.create_sheet(title="Items")
        headers = ["N° Control", "Proveedor", "# Item", "Descripción", "Cantidad", "Precio Unit.", "Total Item"]
        for ci, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=ci, value=h)
            self._style_header_cell(cell)
        row = 2
        for inv in invoices:
            for idx, item in enumerate(inv.items, 1):
                sheet.cell(row=row, column=1, value=inv.control_number or "")
                sheet.cell(row=row, column=2, value=inv.supplier.name)
                sheet.cell(row=row, column=3, value=item.item_number or idx)
                sheet.cell(row=row, column=4, value=item.description)
                sheet.cell(row=row, column=5, value=float(item.quantity))
                c6 = sheet.cell(row=row, column=6, value=float(item.unit_price))
                c7 = sheet.cell(row=row, column=7, value=float(item.total))
                c6.number_format = f"{self.currency_symbol}#,##0.00"
                c7.number_format = f"{self.currency_symbol}#,##0.00"
                row += 1
        self._auto_adjust_widths(sheet)

    def _export_csv(
        self, invoices: list[PurchaseInvoice], columns: list[ColumnDef],
        output_path: str, options: dict[str, Any],
    ) -> str:
        """Export to CSV. / Exporta a CSV."""
        invoices = self._sort_by_group(invoices, options.get("group_by", "none"))
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([col.label for col in columns])
            for inv in invoices:
                writer.writerow([self._get_column_value(inv, col.id) for col in columns])
        logger.info("Exported %d invoices to CSV: %s", len(invoices), output_path)
        return output_path

    def _export_pdf(
        self, invoices: list[PurchaseInvoice], columns: list[ColumnDef],
        output_path: str, options: dict[str, Any],
    ) -> str:
        """Export to PDF with table. / Exporta a PDF con tabla."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, letter
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

        doc = SimpleDocTemplate(
            output_path, pagesize=landscape(letter),
            rightMargin=0.4 * inch, leftMargin=0.4 * inch,
            topMargin=0.4 * inch, bottomMargin=0.4 * inch,
        )
        data: list[list[str]] = [[col.label for col in columns]]
        for inv in invoices:
            row: list[str] = []
            for col in columns:
                val = self._get_column_value(inv, col.id)
                if col.is_currency and isinstance(val, float):
                    val = f"{self.currency_symbol}{val:,.2f}"
                row.append(str(val) if val is not None else "")
            data.append(row)
        ncols = len(columns) or 1
        col_w = 9.2 * inch / ncols
        table = Table(data, colWidths=[col_w] * ncols)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("FONTSIZE", (0, 1), (-1, -1), 6),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("PADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
        ]))
        doc.build([table])
        logger.info("Exported %d invoices to PDF: %s", len(invoices), output_path)
        return output_path

    def _export_json(
        self, invoices: list[PurchaseInvoice], output_path: str, options: dict[str, Any],
    ) -> str:
        """Export to JSON with ALL fields (ignores profile). / Exporta JSON completo."""
        include_raw = options.get("include_raw_data", False)
        data = {
            "metadata": {
                "exported_at": datetime.utcnow().isoformat(),
                "total_invoices": len(invoices),
                "total_amount": sum(float(inv.total) for inv in invoices),
                "currency": self.currency_symbol,
                "format_version": "1.0",
            },
            "invoices": [self._invoice_to_full_dict(inv, include_raw) for inv in invoices],
        }
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        logger.info("Exported %d invoices to JSON: %s", len(invoices), output_path)
        return output_path

    def _invoice_to_full_dict(self, invoice: PurchaseInvoice, include_raw: bool) -> dict[str, Any]:
        """Convert invoice to full dict. / Convierte factura a dict completo."""
        result: dict[str, Any] = {col.id: self._get_column_value(invoice, col.id) for col in ALL_COLUMNS}
        result["items"] = [
            {
                "item_number": item.item_number, "product_code": item.product_code or "",
                "description": item.description, "quantity": float(item.quantity),
                "unit_price": float(item.unit_price), "discount": float(item.discount),
                "taxable_sale": float(item.taxable_sale), "total": float(item.total),
            }
            for item in invoice.items
        ]
        if include_raw and invoice.raw_data:
            result["raw_data"] = invoice.raw_data
        return result

    def _style_header_cell(self, cell: Any) -> None:
        """Apply blue header style. / Aplica estilo azul de header."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

    def _auto_adjust_widths(self, sheet: Any) -> None:
        """Auto-adjust column widths. / Ajusta anchos automaticamente."""
        for column_cells in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                try:
                    max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_len + 2, 50)

    def _sort_by_group(self, invoices: list[PurchaseInvoice], group_by: str) -> list[PurchaseInvoice]:
        """Sort invoices by group. / Ordena facturas por grupo."""
        if group_by == "supplier":
            return sorted(invoices, key=lambda i: i.supplier.name)
        if group_by == "date":
            return sorted(invoices, key=lambda i: i.issue_date)
        if group_by == "type":
            return sorted(invoices, key=lambda i: i.document_type.value)
        return list(invoices)
