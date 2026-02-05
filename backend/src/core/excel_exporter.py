"""
Excel Exporter / Exportador de Excel
====================================

Export invoice data to Excel format.
Exporta datos de facturas a formato Excel.

This module provides:
Este módulo provee:
- export_to_excel: Export invoices to Excel file
                   Exporta facturas a archivo Excel
- create_summary_sheet: Create summary statistics sheet
                        Crea hoja de resumen estadístico
- format_currency: Format decimal values as currency
                   Formatea valores decimales como moneda
"""

import logging
from decimal import Decimal
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.models.invoice import Invoice, InvoiceType

logger = logging.getLogger(__name__)


def payment_text(condition: Optional[int]) -> str:
    """
    Convert payment condition code to readable text.
    Convierte código de condición de pago a texto legible.

    Args / Argumentos:
        condition: Payment condition code (1=CONTADO, 2=CRÉDITO)
                   Código de condición de pago

    Returns / Retorna:
        Human readable payment text / Texto de pago legible
    """
    if condition == 1:
        return "CONTADO"
    elif condition == 2:
        return "CRÉDITO"
    return ""


class ExcelExporterError(Exception):
    """
    Custom exception for Excel export errors.
    Excepción personalizada para errores de exportación Excel.
    """

    def __init__(
        self,
        message: str,
        output_path: Optional[str] = None,
        original_error: Optional[Exception] = None,
    ) -> None:
        self.message = message
        self.output_path = output_path
        self.original_error = original_error
        super().__init__(self.message)


class ExcelExporter:
    """
    Excel invoice exporter / Exportador de facturas a Excel.

    Creates Excel workbooks with invoice data, including formatting
    and summary statistics.

    Crea libros de Excel con datos de facturas, incluyendo formato
    y estadísticas de resumen.

    Attributes / Atributos:
        currency_symbol: Symbol for currency formatting
                         Símbolo para formato de moneda
        decimal_places: Number of decimal places for currency
                        Número de decimales para moneda
    """

    # Column headers / Encabezados de columna
    HEADERS = [
        "N° Control",
        "N° Documento Interno",
        "Document Number / Código Gen.",
        "Type / Tipo",
        "Issue Date / Fecha",
        "Emission Time / Hora",
        "Payment / Condición",
        "Seller / Vendedor",
        "Customer Name / Cliente",
        "Customer ID / NIT",
        "Customer Doc Type / Tipo Doc",
        "Customer NRC",
        "Customer Address / Dirección",
        "Customer Phone / Teléfono",
        "Customer Email / Correo",
        "Issuer Name / Emisor",
        "Issuer NIT",
        "Issuer NRC",
        "Total Taxable / Gravado",
        "Total Exempt / Exento",
        "Total Non-Subject / No Sujeto",
        "Total Discount / Descuento",
        "Subtotal",
        "Tax / IVA",
        "Total",
        "Total in Words / Letras",
        "Currency / Moneda",
        "Tax Seal / Sello",
        "Source File / Archivo",
    ]

    def __init__(
        self,
        currency_symbol: str = "$",
        decimal_places: int = 2,
    ) -> None:
        """
        Initialize the Excel exporter.
        Inicializa el exportador de Excel.

        Args / Argumentos:
            currency_symbol: Currency symbol to use / Símbolo de moneda a usar
            decimal_places: Decimal places for currency / Decimales para moneda
        """
        self.currency_symbol = currency_symbol
        self.decimal_places = decimal_places
        logger.debug(
            "ExcelExporter initialized with symbol=%s, decimals=%d",
            currency_symbol,
            decimal_places,
        )

    def export_to_excel(
        self,
        invoices: list[Invoice],
        output_path: str,
        include_summary: bool = True,
        include_items: bool = False,
    ) -> str:
        """
        Export invoices to an Excel file.
        Exporta facturas a un archivo Excel.

        Args / Argumentos:
            invoices: List of invoices to export / Lista de facturas a exportar
            output_path: Path for the output Excel file
                         Ruta para el archivo Excel de salida
            include_summary: Whether to include summary sheet
                             Si incluir hoja de resumen
            include_items: Whether to include detailed items sheet
                           Si incluir hoja de ítems detallados

        Returns / Retorna:
            Path to the created Excel file / Ruta al archivo Excel creado

        Raises / Lanza:
            ExcelExporterError: If export fails / Si la exportación falla
        """
        if not invoices:
            raise ExcelExporterError("No invoices provided for export")

        logger.info("Exporting %d invoices to %s", len(invoices), output_path)

        try:
            workbook = Workbook()

            # Create main invoices sheet
            self._create_invoices_sheet(workbook, invoices)

            # Create summary sheet if requested
            if include_summary:
                self.create_summary_sheet(workbook, invoices)

            # Create items detail sheet if requested
            if include_items:
                self._create_items_sheet(workbook, invoices)

            # Ensure output directory exists
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)

            workbook.save(str(output_file))

            logger.info("Successfully exported to %s", output_path)
            return str(output_file)

        except ExcelExporterError:
            raise
        except Exception as e:
            raise ExcelExporterError(
                f"Error exporting to Excel: {e}",
                output_path=output_path,
                original_error=e,
            ) from e

    def create_summary_sheet(
        self,
        workbook: Workbook,
        invoices: list[Invoice],
    ) -> None:
        """
        Create a summary statistics sheet in the workbook.
        Crea una hoja de estadísticas de resumen en el libro.

        Args / Argumentos:
            workbook: The workbook to add the sheet to
                      El libro al que agregar la hoja
            invoices: List of invoices for statistics
                      Lista de facturas para estadísticas
        """
        sheet = workbook.create_sheet(title="Summary - Resumen")

        # Calculate statistics
        total_count = len(invoices)
        total_subtotal = sum((inv.subtotal for inv in invoices), Decimal(0))
        total_tax = sum((inv.tax for inv in invoices), Decimal(0))
        total_amount = sum((inv.total for inv in invoices), Decimal(0))

        # Count by type
        type_counts = dict.fromkeys(InvoiceType, 0)
        for inv in invoices:
            type_counts[inv.invoice_type] += 1

        # Date range
        dates = [inv.issue_date for inv in invoices]
        min_date = min(dates) if dates else None
        max_date = max(dates) if dates else None

        # Write summary data
        summary_data = [
            ("Summary / Resumen", ""),
            ("", ""),
            ("Total Invoices / Total Facturas", total_count),
            ("", ""),
            ("By Type / Por Tipo:", ""),
            ("  Factura", type_counts[InvoiceType.FACTURA]),
            ("  CCF", type_counts[InvoiceType.CCF]),
            ("  Nota Crédito", type_counts[InvoiceType.NOTA_CREDITO]),
            ("", ""),
            ("Date Range / Rango de Fechas:", ""),
            ("  From / Desde", min_date.isoformat() if min_date else "N/A"),
            ("  To / Hasta", max_date.isoformat() if max_date else "N/A"),
            ("", ""),
            ("Totals / Totales:", ""),
            ("  Subtotal", self.format_currency(total_subtotal)),
            ("  Tax / Impuesto", self.format_currency(total_tax)),
            ("  Grand Total / Total General", self.format_currency(total_amount)),
        ]

        for row_num, (label, value) in enumerate(summary_data, start=1):
            sheet.cell(row=row_num, column=1, value=label)
            sheet.cell(row=row_num, column=2, value=value)

        # Apply styling
        self._style_summary_sheet(sheet)

        logger.debug("Created summary sheet with %d invoices", total_count)

    def format_currency(self, value: Decimal) -> str:
        """
        Format a decimal value as currency.
        Formatea un valor decimal como moneda.

        Args / Argumentos:
            value: Decimal value to format / Valor decimal a formatear

        Returns / Retorna:
            Formatted currency string / String de moneda formateado
        """
        format_str = f"{{symbol}}{{value:,.{self.decimal_places}f}}"
        return format_str.format(symbol=self.currency_symbol, value=float(value))

    def _create_invoices_sheet(
        self,
        workbook: Workbook,
        invoices: list[Invoice],
    ) -> None:
        """
        Create the main invoices data sheet with ALL DTE fields.
        Crea la hoja principal de datos de facturas con TODOS los campos DTE.
        """
        sheet = workbook.active
        sheet.title = "Invoices - Facturas"

        # Write headers
        for col, header in enumerate(self.HEADERS, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._style_header_cell(cell)

        # Write invoice data with ALL fields
        for row_num, invoice in enumerate(invoices, start=2):
            col = 1
            sheet.cell(row=row_num, column=col, value=invoice.control_number or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.internal_doc_number or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.document_number); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.invoice_type.value); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.issue_date.isoformat()); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.emission_time or ""); col += 1
            sheet.cell(row=row_num, column=col, value=payment_text(invoice.payment_condition)); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.seller_name or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_name); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_id or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_doc_type or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_nrc or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_address or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_phone or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.customer_email or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.issuer_name or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.issuer_nit or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.issuer_nrc or ""); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.total_taxable)); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.total_exempt)); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.total_non_subject)); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.total_discount)); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.subtotal)); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.tax)); col += 1
            sheet.cell(row=row_num, column=col, value=float(invoice.total)); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.total_in_words or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.currency); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.tax_seal or ""); col += 1
            sheet.cell(row=row_num, column=col, value=invoice.source_file or "")

            # Apply currency format to numeric cells (columns 19-25)
            for currency_col in [19, 20, 21, 22, 23, 24, 25]:
                cell = sheet.cell(row=row_num, column=currency_col)
                cell.number_format = f"{self.currency_symbol}#,##0.00"

        # Auto-adjust column widths
        self._auto_adjust_columns(sheet)

        logger.debug("Created invoices sheet with %d rows", len(invoices))

    def _create_items_sheet(
        self,
        workbook: Workbook,
        invoices: list[Invoice],
    ) -> None:
        """
        Create a detailed items sheet with ALL DTE fields.
        Crea una hoja de ítems detallados con TODOS los campos DTE.
        """
        sheet = workbook.create_sheet(title="Items - Items")

        item_headers = [
            "Invoice / Factura",
            "N° Control",
            "Item # / Ítem #",
            "Product Code / Código",
            "Description / Descripción",
            "Unit Measure / Unidad",
            "Quantity / Cantidad",
            "Original Price / Precio Orig.",
            "Unit Price / Precio Unit.",
            "Discount / Descuento",
            "Taxable / Gravada",
            "Exempt / Exenta",
            "Non-Subject / No Sujeta",
            "Item Tax / IVA Ítem",
            "Total",
        ]

        # Write headers
        for col, header in enumerate(item_headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._style_header_cell(cell)

        # Write item data with ALL DTE fields
        row_num = 2
        for invoice in invoices:
            for idx, item in enumerate(invoice.items, start=1):
                col = 1
                sheet.cell(row=row_num, column=col, value=invoice.document_number); col += 1
                sheet.cell(row=row_num, column=col, value=invoice.control_number or ""); col += 1
                sheet.cell(row=row_num, column=col, value=item.item_number or idx); col += 1
                sheet.cell(row=row_num, column=col, value=item.product_code or ""); col += 1
                sheet.cell(row=row_num, column=col, value=item.description); col += 1
                sheet.cell(row=row_num, column=col, value=item.unit_measure or ""); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.quantity)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.original_price) if item.original_price else ""); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.unit_price)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.discount)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.taxable_sale)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.exempt_sale)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.non_subject_sale)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.item_tax)); col += 1
                sheet.cell(row=row_num, column=col, value=float(item.total))

                # Apply currency format to money columns (8-15)
                for currency_col in [8, 9, 10, 11, 12, 13, 14, 15]:
                    cell = sheet.cell(row=row_num, column=currency_col)
                    if cell.value != "":
                        cell.number_format = f"{self.currency_symbol}#,##0.00"

                row_num += 1

        self._auto_adjust_columns(sheet)

        logger.debug("Created items sheet with ALL DTE fields")

    def _style_header_cell(self, cell) -> None:
        """
        Apply header styling to a cell.
        Aplica estilo de encabezado a una celda.
        """
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def _style_summary_sheet(self, sheet) -> None:
        """
        Apply styling to the summary sheet.
        Aplica estilo a la hoja de resumen.
        """
        # Title styling
        title_cell = sheet.cell(row=1, column=1)
        title_cell.font = Font(bold=True, size=14)

        # Section headers
        for row in [5, 10, 14]:
            cell = sheet.cell(row=row, column=1)
            cell.font = Font(bold=True)

        # Adjust column widths
        sheet.column_dimensions["A"].width = 35
        sheet.column_dimensions["B"].width = 20

    def _auto_adjust_columns(self, sheet) -> None:
        """
        Auto-adjust column widths based on content.
        Ajusta automáticamente anchos de columna según contenido.
        """
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def export_to_csv(
        self,
        invoices: list[Invoice],
        output_path: str,
        include_header: bool = True,
        delimiter: str = ",",
    ) -> str:
        """
        Export invoices to CSV format with ALL DTE fields.
        Exporta facturas a formato CSV con TODOS los campos DTE.

        Args / Argumentos:
            invoices: List of invoices to export / Lista de facturas a exportar
            output_path: Path for output file / Ruta del archivo de salida
            include_header: Include column headers / Incluir encabezados
            delimiter: Field delimiter / Delimitador de campos

        Returns / Retorna:
            Path to created file / Ruta al archivo creado
        """
        import csv

        if not invoices:
            raise ExcelExporterError("No invoices provided for export")

        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".csv")

        output_file.parent.mkdir(parents=True, exist_ok=True)

        with open(output_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f, delimiter=delimiter)

            if include_header:
                writer.writerow(self.HEADERS)

            for invoice in invoices:
                writer.writerow([
                    invoice.control_number or "",
                    invoice.internal_doc_number or "",
                    invoice.document_number,
                    invoice.invoice_type.value,
                    invoice.issue_date.isoformat(),
                    invoice.emission_time or "",
                    payment_text(invoice.payment_condition),
                    invoice.seller_name or "",
                    invoice.customer_name,
                    invoice.customer_id or "",
                    invoice.customer_doc_type or "",
                    invoice.customer_nrc or "",
                    invoice.customer_address or "",
                    invoice.customer_phone or "",
                    invoice.customer_email or "",
                    invoice.issuer_name or "",
                    invoice.issuer_nit or "",
                    invoice.issuer_nrc or "",
                    float(invoice.total_taxable),
                    float(invoice.total_exempt),
                    float(invoice.total_non_subject),
                    float(invoice.total_discount),
                    float(invoice.subtotal),
                    float(invoice.tax),
                    float(invoice.total),
                    invoice.total_in_words or "",
                    invoice.currency,
                    invoice.tax_seal or "",
                    invoice.source_file or "",
                ])

        logger.info("Exported %d invoices to CSV: %s", len(invoices), output_file)
        return str(output_file)

    def export_to_json(
        self,
        invoices: list[Invoice],
        output_path: str,
        indent: int = 2,
        include_metadata: bool = True,
    ) -> str:
        """
        Export invoices to JSON format with ALL DTE fields.
        Exporta facturas a formato JSON con TODOS los campos DTE.

        Args / Argumentos:
            invoices: List of invoices to export / Lista de facturas a exportar
            output_path: Path for output file / Ruta del archivo de salida
            indent: JSON indentation / Indentación del JSON
            include_metadata: Include export metadata / Incluir metadata de exportación

        Returns / Retorna:
            Path to created file / Ruta al archivo creado
        """
        import json
        from datetime import datetime

        if not invoices:
            raise ExcelExporterError("No invoices provided for export")

        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".json")

        output_file.parent.mkdir(parents=True, exist_ok=True)

        def invoice_to_dict(inv: Invoice) -> dict:
            return {
                # Core identification / Identificación
                "document_number": inv.document_number,
                "control_number": inv.control_number or "",
                "internal_doc_number": inv.internal_doc_number or "",
                "type": inv.invoice_type.value,
                "issue_date": inv.issue_date.isoformat(),
                "emission_time": inv.emission_time or "",
                "currency": inv.currency,

                # Issuer / Emisor
                "issuer": {
                    "name": inv.issuer_name or "",
                    "commercial_name": inv.issuer_commercial_name or "",
                    "nit": inv.issuer_nit or "",
                    "nrc": inv.issuer_nrc or "",
                    "address": inv.issuer_address or "",
                },

                # Customer / Cliente
                "customer": {
                    "name": inv.customer_name,
                    "id": inv.customer_id or "",
                    "doc_type": inv.customer_doc_type or "",
                    "nrc": inv.customer_nrc or "",
                    "address": inv.customer_address or "",
                    "phone": inv.customer_phone or "",
                    "email": inv.customer_email or "",
                },

                # Amounts / Montos
                "amounts": {
                    "subtotal": float(inv.subtotal),
                    "tax": float(inv.tax),
                    "total": float(inv.total),
                    "total_taxable": float(inv.total_taxable),
                    "total_exempt": float(inv.total_exempt),
                    "total_non_subject": float(inv.total_non_subject),
                    "total_discount": float(inv.total_discount),
                    "total_in_words": inv.total_in_words or "",
                },

                # Payment / Pago
                "payment_condition": inv.payment_condition,
                "seller_name": inv.seller_name or "",

                # Items / Ítems
                "items": [
                    {
                        "item_number": item.item_number,
                        "product_code": item.product_code or "",
                        "description": item.description,
                        "unit_measure": item.unit_measure,
                        "quantity": float(item.quantity),
                        "original_price": float(item.original_price) if item.original_price else None,
                        "unit_price": float(item.unit_price),
                        "discount": float(item.discount),
                        "taxable_sale": float(item.taxable_sale),
                        "exempt_sale": float(item.exempt_sale),
                        "non_subject_sale": float(item.non_subject_sale),
                        "item_tax": float(item.item_tax),
                        "total": float(item.total),
                    }
                    for item in inv.items
                ],

                # Tax seal / Sello fiscal
                "tax_seal": inv.tax_seal or "",
                "source_file": inv.source_file or "",
            }

        # Always include base structure with metadata first for consistency
        # Siempre incluir estructura base con metadata primero para consistencia
        total_amount = sum(float(inv.total) for inv in invoices)

        data: dict = {
            "metadata": {
                "exported_at": datetime.utcnow().isoformat(),
                "total_invoices": len(invoices),
                "total_amount": total_amount,
                "currency": self.currency_symbol,
                "format_version": "1.0",
            },
            "invoices": [invoice_to_dict(inv) for inv in invoices],
        }

        # Remove detailed metadata if not requested, but keep minimal structure
        # Remover metadata detallada si no se solicita, pero mantener estructura mínima
        if not include_metadata:
            data["metadata"] = {
                "total_invoices": len(invoices),
                "format_version": "1.0",
            }

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)

        logger.info("Exported %d invoices to JSON: %s", len(invoices), output_file)
        return str(output_file)

    def export_to_pdf(
        self,
        invoices: list[Invoice],
        output_path: str,
        include_summary: bool = True,
        title: str = "Invoice Report / Reporte de Facturas",
    ) -> str:
        """
        Export invoices to PDF format with one invoice per page.
        Exporta facturas a formato PDF con una factura por página.

        Args / Argumentos:
            invoices: List of invoices to export / Lista de facturas a exportar
            output_path: Path for output file / Ruta del archivo de salida
            include_summary: Include summary section / Incluir sección de resumen
            title: Report title / Título del reporte

        Returns / Retorna:
            Path to created file / Ruta al archivo creado
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        if not invoices:
            raise ExcelExporterError("No invoices provided for export")

        output_file = Path(output_path)
        if not output_file.suffix:
            output_file = output_file.with_suffix(".pdf")

        output_file.parent.mkdir(parents=True, exist_ok=True)

        doc = SimpleDocTemplate(
            str(output_file),
            pagesize=letter,
            rightMargin=0.5 * inch,
            leftMargin=0.5 * inch,
            topMargin=0.5 * inch,
            bottomMargin=0.5 * inch,
        )

        elements = []
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=20,
            alignment=1,  # Center
        )
        section_style = ParagraphStyle(
            "SectionHeader",
            parent=styles["Heading2"],
            fontSize=12,
            spaceBefore=10,
            spaceAfter=5,
            textColor=colors.HexColor("#4472C4"),
        )
        label_style = ParagraphStyle(
            "Label",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.grey,
        )

        # ========== SUMMARY PAGE ==========
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 20))

        if include_summary:
            total_amount = sum(float(inv.total) for inv in invoices)
            total_taxable = sum(float(inv.total_taxable) for inv in invoices)
            total_exempt = sum(float(inv.total_exempt) for inv in invoices)
            total_tax = sum(float(inv.tax) for inv in invoices)

            # Summary table
            summary_data = [
                ["RESUMEN GENERAL", ""],
                ["Total de Facturas", str(len(invoices))],
                ["Total Gravado", f"{self.currency_symbol}{total_taxable:,.2f}"],
                ["Total Exento", f"{self.currency_symbol}{total_exempt:,.2f}"],
                ["Total IVA", f"{self.currency_symbol}{total_tax:,.2f}"],
                ["TOTAL GENERAL", f"{self.currency_symbol}{total_amount:,.2f}"],
            ]
            summary_table = Table(summary_data, colWidths=[3.5 * inch, 2.5 * inch])
            summary_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F0FE")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#CCCCCC")),
                ("FONTSIZE", (0, 1), (-1, -1), 10),
                ("PADDING", (0, 0), (-1, -1), 8),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ]))
            elements.append(summary_table)
            elements.append(Spacer(1, 30))

            # Invoice list table (compact overview)
            list_headers = ["#", "Fecha", "Cliente", "Tipo", "Total"]
            list_data = [list_headers]
            for i, inv in enumerate(invoices, 1):
                list_data.append([
                    str(i),
                    inv.issue_date.strftime("%Y-%m-%d"),
                    (inv.customer_name or "-")[:35],
                    inv.invoice_type.value,
                    f"{self.currency_symbol}{float(inv.total):,.2f}",
                ])

            list_table = Table(
                list_data,
                colWidths=[0.4 * inch, 1 * inch, 3 * inch, 1 * inch, 1.2 * inch]
            )
            list_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("PADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (0, 0), (0, -1), "CENTER"),
                ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
            ]))
            elements.append(Paragraph("Listado de Facturas", section_style))
            elements.append(list_table)

        # ========== INDIVIDUAL INVOICE PAGES (COMPACT LAYOUT) ==========
        # Compact style for fitting everything on one page
        compact_style = TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("PADDING", (0, 0), (-1, -1), 3),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])

        for idx, inv in enumerate(invoices):
            elements.append(PageBreak())

            # Invoice header - compact
            header_style = ParagraphStyle(
                "CompactTitle",
                parent=styles["Heading1"],
                fontSize=14,
                spaceAfter=8,
                alignment=1,
            )
            elements.append(Paragraph(
                f"FACTURA {idx + 1} de {len(invoices)}",
                header_style
            ))

            # ===== ROW 1: Document Info (left) + Amounts (right) =====
            # Document info - compact 4 rows
            doc_data = [
                ["DOCUMENTO", "", "", ""],
                ["Código", (inv.document_number or "-")[:25], "Control", inv.control_number or "-"],
                ["Fecha", f"{inv.issue_date.strftime('%Y-%m-%d')} {inv.emission_time or ''}", "Tipo", inv.invoice_type.value],
                ["Pago", payment_text(inv.payment_condition) or "-", "Moneda", inv.currency or "USD"],
            ]
            doc_table = Table(doc_data, colWidths=[0.6 * inch, 1.5 * inch, 0.55 * inch, 1.0 * inch])
            doc_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("SPAN", (0, 0), (-1, 0)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0F0F0")),
                ("BACKGROUND", (2, 1), (2, -1), colors.HexColor("#F0F0F0")),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTNAME", (2, 1), (2, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("PADDING", (0, 0), (-1, -1), 3),
            ]))

            # Amounts - compact
            amt_data = [
                ["MONTOS", ""],
                ["Gravado", f"{self.currency_symbol}{float(inv.total_taxable):,.2f}"],
                ["Exento", f"{self.currency_symbol}{float(inv.total_exempt):,.2f}"],
                ["No Sujeto", f"{self.currency_symbol}{float(inv.total_non_subject):,.2f}"],
                ["Descuento", f"{self.currency_symbol}{float(inv.total_discount):,.2f}"],
                ["Subtotal", f"{self.currency_symbol}{float(inv.subtotal):,.2f}"],
                ["IVA", f"{self.currency_symbol}{float(inv.tax):,.2f}"],
                ["TOTAL", f"{self.currency_symbol}{float(inv.total):,.2f}"],
            ]
            amt_table = Table(amt_data, colWidths=[1.0 * inch, 1.2 * inch])
            amt_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("SPAN", (0, 0), (-1, 0)),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8F0FE")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, -1), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("BACKGROUND", (0, 1), (0, -2), colors.HexColor("#F0F0F0")),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -2), 7),
                ("PADDING", (0, 0), (-1, -1), 3),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
            ]))

            row1 = Table([[doc_table, amt_table]], colWidths=[3.8 * inch, 2.4 * inch])
            row1.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            elements.append(row1)
            elements.append(Spacer(1, 6))

            # ===== ROW 2: Issuer (left) + Customer (right) =====
            issuer_data = [
                ["EMISOR", ""],
                ["Nombre", (inv.issuer_name or "-")[:30]],
                ["NIT", inv.issuer_nit or "-"],
                ["NRC", inv.issuer_nrc or "-"],
            ]
            issuer_table = Table(issuer_data, colWidths=[0.7 * inch, 2.3 * inch])
            issuer_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("SPAN", (0, 0), (-1, 0)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0F0F0")),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("PADDING", (0, 0), (-1, -1), 3),
            ]))

            customer_data = [
                ["CLIENTE", ""],
                ["Nombre", (inv.customer_name or "-")[:30]],
                ["NIT/DUI", inv.customer_id or "-"],
                ["NRC", inv.customer_nrc or "-"],
                ["Dirección", (inv.customer_address or "-")[:35]],
                ["Tel/Email", f"{inv.customer_phone or '-'} / {(inv.customer_email or '-')[:20]}"],
            ]
            customer_table = Table(customer_data, colWidths=[0.7 * inch, 2.5 * inch])
            customer_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("SPAN", (0, 0), (-1, 0)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0F0F0")),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("PADDING", (0, 0), (-1, -1), 3),
            ]))

            row2 = Table([[issuer_table, customer_table]], colWidths=[3.2 * inch, 3.4 * inch])
            row2.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
            elements.append(row2)
            elements.append(Spacer(1, 6))

            # ===== ROW 3: Total in words + Additional info =====
            extra_data = [
                ["INFORMACIÓN ADICIONAL", "", ""],
                ["Total en Letras", (inv.total_in_words or "-")[:50], ""],
                ["Vendedor", inv.seller_name or "-", ""],
                ["Sello", (inv.tax_seal or "-")[:45], ""],
                ["Archivo", (inv.source_file or "-")[:40], ""],
            ]
            extra_table = Table(extra_data, colWidths=[1.0 * inch, 3.5 * inch, 2.0 * inch])
            extra_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("SPAN", (0, 0), (-1, 0)),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#F0F0F0")),
                ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("PADDING", (0, 0), (-1, -1), 3),
            ]))
            elements.append(extra_table)

        doc.build(elements)
        logger.info("Exported %d invoices to PDF: %s", len(invoices), output_file)
        return str(output_file)
